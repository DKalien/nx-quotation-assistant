"""
数据导出器

将提取的模型数据导出为各种格式用于报价。
"""

import csv
import json
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Font = None
    Alignment = None
    Border = None
    Side = None


class DataExporter:
    """
    将模型数据导出为各种格式。

    支持 CSV、Excel、JSON 导出格式。
    """

    def __init__(self, output_dir='.'):
        """
        初始化导出器。

        参数:
            output_dir: 输出文件的目录
        """
        self.output_dir = output_dir

    def to_csv(self, data, filename):
        """
        将数据导出到 CSV 文件。

        参数:
            data: 包含模型数据的字典列表
            filename: 输出文件名

        返回:
            str: 创建的文件路径
        """
        if not data:
            return None

        filepath = f"{self.output_dir}/{filename}"

        # 获取所有唯一的键作为表头
        if isinstance(data, list):
            keys = set()
            for item in data:
                keys.update(item.keys())
            keys = sorted(list(keys))
        else:
            keys = list(data.keys())
            data = [data]

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(data)

        return filepath

    def to_json(self, data, filename):
        """
        将数据导出到 JSON 文件。

        参数:
            data: 字典或字典列表
            filename: 输出文件名

        返回:
            str: 创建的文件路径
        """
        filepath = f"{self.output_dir}/{filename}"

        # 将元组转换为列表以便 JSON 序列化
        def convert(obj):
            if isinstance(obj, dict):
                return {k: convert(v) for k, v in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [convert(item) for item in obj]
            else:
                return obj

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(convert(data), f, indent=2, ensure_ascii=False)

        return filepath

    def to_excel(self, data, filename):
        """
        将数据导出到 Excel 文件。

        需要 openpyxl 包。

        参数:
            data: 包含模型数据的字典列表
            filename: 输出文件名

        返回:
            str: 创建的文件路径，如果 openpyxl 不可用则返回 None
        """
        if not OPENPYXL_AVAILABLE:
            print("未安装 openpyxl。运行: pip install openpyxl")
            return None

        if not data:
            return None

        filepath = f"{self.output_dir}/{filename}"

        wb = Workbook()
        ws = wb.active
        ws.title = "报价数据"

        # 将单个字典转换为列表
        if isinstance(data, dict):
            data = [data]

        # 获取表头
        headers = list(data[0].keys()) if data else []

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 写入数据
        for row_idx, item in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                value = item.get(key, '')
                # 将元组转换为字符串
                if isinstance(value, (list, tuple)):
                    value = str(value)
                ws.cell(row=row_idx, column=col, value=value)

        # 添加元数据
        ws_meta = wb.create_sheet("元数据")
        ws_meta['A1'] = "由 NX 报价助手生成"
        ws_meta['A2'] = f"日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_meta['A3'] = f"零件总数: {len(data)}"

        wb.save(filepath)

        return filepath

    def create_quotation_report(self, data, filename):
        """
        创建格式化的报价报告。

        参数:
            data: 模型数据字典
            filename: 输出文件名

        返回:
            str: 创建的文件路径，如果 openpyxl 不可用则返回 JSON 文件路径
        """
        if not OPENPYXL_AVAILABLE:
            return self.to_json(data, filename.replace('.xlsx', '.json'))

        filepath = f"{self.output_dir}/{filename}"
        wb = Workbook()
        ws = wb.active
        ws.title = "报价报告"

        # 样式
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 标题
        ws['A1'] = "模型参数报告"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

        # 表头
        headers = ['参数', '值', '单位']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # 数据行
        row = 5
        params = [
            ('零件名称', data.get('part_name', ''), ''),
            ('质量', data.get('mass', 0), 'kg'),
            ('体积', data.get('volume', 0), 'mm³'),
            ('表面积', data.get('surface_area', 0), 'mm²'),
            ('长度', data.get('length', 0), 'mm'),
            ('宽度', data.get('width', 0), 'mm'),
            ('高度', data.get('height', 0), 'mm'),
            ('实体数量', data.get('body_count', 0), ''),
        ]

        for param, value, unit in params:
            ws.cell(row=row, column=1, value=param).border = border
            cell_val = ws.cell(row=row, column=2, value=value)
            cell_val.border = border
            if isinstance(value, float):
                cell_val.number_format = '0.0000'
            ws.cell(row=row, column=3, value=unit).border = border
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10

        wb.save(filepath)
        return filepath
